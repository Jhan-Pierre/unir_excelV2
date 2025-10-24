import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO

st.title("📊 Consolidador de Archivos Excel (.xlsm / .xlsx)")

st.markdown("""
Esta aplicación consolida datos de varios archivos `.xlsm` o `.xlsx` en una **plantilla base**, 
manteniendo las fórmulas y formatos existentes en las hojas de la plantilla.
""")

# --- Carga de archivos ---
uploaded_files = st.file_uploader(
    "Sube los archivos Excel (.xlsm o .xlsx) a consolidar",
    type=["xlsm", "xlsx"],  # 🔹 Aceptar ambos formatos
    accept_multiple_files=True
)

template_file = st.file_uploader(
    "Sube el archivo plantilla (.xlsm o .xlsx)",
    type=["xlsm", "xlsx"]  # 🔹 Aceptar ambos formatos
)

if uploaded_files and template_file:
    st.success(f"Se cargaron {len(uploaded_files)} archivos y 1 plantilla.")
    
    if st.button("🔄 Consolidar"):
        
        # --- NUEVO: Elementos de progreso ---
        progress_bar = st.progress(0)
        status_text = st.empty()
        # ------------------------------------
        
        try:
            # Cargar plantilla (mantiene macros si las hay)
            plantilla = load_workbook(template_file, keep_vba=True)
            
            total_sheets = len(plantilla.sheetnames)
            total_files = len(uploaded_files)
            total_steps = total_sheets * total_files
            steps_done = 0

            if total_steps == 0:
                st.warning("No hay archivos o hojas para procesar.")
                progress_bar.empty()
                status_text.empty()
                st.stop()

            for i, hoja in enumerate(plantilla.sheetnames):
                ws_plantilla = plantilla[hoja]
                start_row = 2
                current_row = start_row
                
                for j, uploaded in enumerate(uploaded_files):
                    steps_done += 1
                    progress_percentage = steps_done / total_steps
                    progress_bar.progress(progress_percentage)
                    status_text.text(
                        f"Procesando: Hoja '{hoja}' ({i+1}/{total_sheets}) | "
                        f"Archivo '{uploaded.name}' ({j+1}/{total_files})"
                    )
                    
                    try:
                        xls = pd.ExcelFile(uploaded)
                        if hoja in xls.sheet_names:
                            df = pd.read_excel(xls, sheet_name=hoja, header=0)
                            for _, row in df.iterrows():
                                current_row += 1
                                for col_idx, value in enumerate(row, start=1):
                                    ws_plantilla.cell(row=current_row, column=col_idx, value=value)
                    except Exception as e:
                        st.warning(f"Error al procesar '{uploaded.name}' (Hoja: '{hoja}'): {e}")
            
            status_text.text("Proceso completado. Guardando archivo final...")
            progress_bar.progress(1.0)

            output = BytesIO()
            plantilla.save(output)
            output.seek(0)
            
            progress_bar.empty()
            status_text.success("Consolidación completada con éxito ✅")
            
            st.download_button(
                label="⬇️ Descargar archivo consolidado",
                data=output,
                file_name="consolidado.xlsm",
                mime="application/vnd.ms-excel.sheet.macroEnabled.12"
            )
        
        except Exception as e:
            st.error(f"Ocurrió un error general: {e}")
            progress_bar.empty()
            status_text.empty()

else:
    st.info("Por favor, carga los archivos a consolidar y la plantilla base para continuar.")
